import unittest
from io import BytesIO
from user3 import main
import pandas as pd
from unittest.mock import Mock,patch
import zipfile
import os
from openpyxl import Workbook

class TestExcelFileExtraction(unittest.TestCase):
    #Verify Excel file Download
    @patch("user3.requests.get")
    def test_case1(self,mock_get):
        #create a fake excel sheet
        wb=Workbook()
        ws=wb.active
        ws['a1']="hello"
        ws['a2']="hitesh"
        ws['b1']=100
        ws["b2"]=200

        excel_bytes=BytesIO() #fake Bytes 
        wb.save(excel_bytes)

        zip_bytes=BytesIO()
        with zipfile.ZipFile(zip_bytes,"w") as z_f:
            z_f.writestr("output_mock.xlsx",excel_bytes.getvalue()) #creating a fake zipe file with excel file stored in it


        #mock the url with the content of fake zip file url
        mock_response=Mock()
        mock_response.status_code=200
        mock_response.content=zip_bytes.getvalue()
        mock_get.return_value=mock_response

        df,file_path=main()

        #chcek if dataframe returns something and the file path exists
        self.assertIsNotNone(df)
        self.assertTrue(os.path.exists(file_path))
        
    #verify Excel file extraction
    @patch("user3.requests.get")
    def test_case2(self,mock_get):
        wb=Workbook()
        ws=wb.active
        ws['a1']="hello"
        ws['a2']="hitesh"
        ws['b1']=100
        ws['b2']=200

        excel_bytes=BytesIO()
        wb.save(excel_bytes)

        zip_bytes=BytesIO()
        with zipfile.ZipFile(zip_bytes,"w") as z_f:
            z_f.writestr("output_mock.xlsx",excel_bytes.getvalue())

        mock_response = Mock()
        mock_response.status_code = 200
        mock_response.content = zip_bytes.getvalue()
        mock_get.return_value = mock_response

        df, file_path = main()

        #check if file path ends is correct,check the length of rows and columns are greater than one
        self.assertTrue(file_path.endswith(".xlsx"))
        self.assertGreater(len(df.columns),0)
        self.assertGreater(len(df),0)

    #validate the file type and format
    @patch("user3.requests.get")
    def test_case3(self, mock_get):
        wb = Workbook()
        ws = wb.active
        ws['a1'] = "hello"
        ws['a2'] = "hitesh"
        ws['b1'] = 100
        ws["b2"] = 200

        excel_bytes = BytesIO()
        wb.save(excel_bytes)

        zip_bytes = BytesIO()
        with zipfile.ZipFile(zip_bytes, "w") as z_f:
            z_f.writestr("output_mock.xlsx", excel_bytes.getvalue())

        mock_response = Mock()
        mock_response.status_code = 200
        mock_response.content = zip_bytes.getvalue()
        mock_get.return_value = mock_response

        df, file_path = main()

        #check file type , if the path exists, and also check if the dataframe returned or not
        self.assertTrue(file_path.endswith(".xlsx"))
        self.assertTrue(os.path.exists(file_path))
        self.assertIsInstance(df,pd.DataFrame)

    #handle Data Structure
    @patch("user3.requests.get")
    def test_case4(self, mock_get):
        #create a fake dataframe to check if all the expected columns exists or not
        df_fake=pd.DataFrame({"Employee ID":["E1011"],"Full Name":["Hitesh"],"Job Title":["Manager"],"Department":["IT"],
                              "Business Unit":["xYZ"],"Gender":["male"],"Ethnicity":["asian"],"Age":[19],"Hire Date":["19-12-2005"],
                              "Annual Salary":[20000],"Bonus":[5000],"Country":["india"],"City":["mangalore"],
                              "Exit Date":["19-12-2024"]})

        excel_bytes = BytesIO()
        df_fake.to_excel(excel_bytes) #convert to excel
        excel_bytes.seek(0)
    

        zip_bytes = BytesIO()
        with zipfile.ZipFile(zip_bytes, "w") as z_f:
            z_f.writestr("output_mock.xlsx", excel_bytes.getvalue())
        zip_bytes.seek(0)

        mock_response = Mock()
        mock_response.status_code = 200
        mock_response.content = zip_bytes.getvalue()
        mock_get.return_value = mock_response

        df, file_path = main()

        #check if all the required columns exists or not
        expected_columns=["Employee ID","Full Name","Job Title","Hire Date"]

        for col in expected_columns:
            self.assertIn(col,df.columns)
    
    #Handle missing and invalid data
    @patch("user3.requests.get")
    def test_case5(self, mock_get):
        #same create dataframe to check if any columns are empty
        df_fake = pd.DataFrame(
            {"Employee ID": ["E1011"], "Full Name": ["Hitesh"], "Job Title": ["Manager"], "Department": ["IT"],
             "Business Unit": ["xYZ"], "Gender": ["male"], "Ethnicity": ["asian"], "Age": [19],
             "Hire Date": ["19-12-2005"],
             "Annual Salary": [20000], "Bonus": [5000], "Country": ["india"], "City": ["mangalore"],
             "Exit Date": ["19-12-2024"]})
        

        excel_bytes = BytesIO()
        df_fake.to_excel(excel_bytes)
        excel_bytes.seek(0)

        zip_bytes = BytesIO()
        with zipfile.ZipFile(zip_bytes, "w") as z_f:
            z_f.writestr("output_mock.xlsx", excel_bytes.getvalue())
        zip_bytes.seek(0)

        mock_response = Mock()
        mock_response.status_code = 200
        mock_response.content = zip_bytes.getvalue()
        mock_get.return_value = mock_response

        df, file_path = main()

        expected_columns = ["Employee ID", "Full Name", "Job Title", "Hire Date"]

        #chcek if any column is having missing values and also null values
        for col in expected_columns:
            self.assertIn(col, df.columns,"columns is missing")

        self.assertFalse(df.isnull().any().any() ,"no null values Exists")

if __name__=="__main__":
    unittest.main()